import os
import shutil
import time

from openpyxl.workbook import Workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as ex
from openpyxl import load_workbook


def save_postanovy_data(driver: webdriver.Firefox, vp_num_value, secret_num_value):
    postanovy_table = WebDriverWait(driver, 4).until(
        ex.visibility_of_element_located((By.XPATH, '/html/body/div[1]/ui-view/section[1]/div[2]/div/table[3]/tbody')))
    post_res = doc_data(postanovy_table, driver, doc_type='Перелік постанов', vp_num_value=vp_num_value,
                        secret_num=secret_num_value)
    if post_res:
        print('шукаємо інші документи')
        other_docs = WebDriverWait(driver, 10).until(
            ex.element_to_be_clickable((By.XPATH, "//div[contains(text(), 'Інші документи')]"))
        )
        print('натискаємо інші документи')
        other_docs.click()

        other_doc_table = WebDriverWait(driver, 4).until(
            ex.visibility_of_element_located(
                (By.XPATH, '/html/body/div[1]/ui-view/section[1]/div[2]/div/table[4]/tbody')))
        doc_data(other_doc_table, driver, doc_type='Інші документи', vp_num_value=vp_num_value,
                 secret_num=secret_num_value)

    # for column, values in postanovy_names.items():
    #     document_name = values['document_name']
    #     if document_name in driver.page_source:
    #         element = WebDriverWait(driver, 4).until(
    #             EC.visibility_of_all_elements_located((By.XPATH,
    #                                                    f"//tr[@data-ng-repeat='item in vm.data.vpView.documents' and contains(., '{document_name}')]"))
    #         )
    #         last_document = element[-1]
    #         last_date = last_document.find_elements(By.TAG_NAME, 'td')[-3].text
    #         for doc in element:
    #             last_document = doc
    #
    #             success_pdf_switch = open_and_switch_to_pdf(driver, last_document)
    #
    #             element = WebDriverWait(driver, 10).until(
    #                 EC.element_to_be_clickable(
    #                     (By.XPATH,
    #                      "//a[@data-ng-click='vm.events.btnViewFileDownloadFile()']"))
    #             )
    #             element.click()
    #
    #             button_back = WebDriverWait(driver, 10).until(
    #                 EC.element_to_be_clickable(
    #                     (By.XPATH, "//button[contains(text(), 'Назад')]"))
    #             )
    #             button_back.click()
    #     else:
    #         print(f'no document {document_name} on page source')


def doc_data(postanovy_table, driver, doc_type, vp_num_value, secret_num):
    all_postanovas = postanovy_table.find_elements(By.TAG_NAME, 'tr')
    wb = create_or_take_workbook()
    sheet = wb.active
    for postanova in all_postanovas:
        # try:
        postanova_data = WebDriverWait(postanova, 10).until(
            ex.visibility_of_all_elements_located((By.TAG_NAME, 'td'))
        )
        doc_number = postanova_data[0].text
        doc_date = postanova_data[1].text
        doc_name = postanova_data[2].text
        format_small = '.pdf'
        format_high = '.PDF'
        if doc_name.endswith('.zip') or doc_name.endswith('.ZIP'):
            format_small = '.zip'
            format_high = '.ZIP'
        elif doc_name.endswith('.html') or doc_name.endswith('.HTML'):
            format_small = '.html'
            format_high = '.HTML'

        new_file_name = f"{doc_number} {doc_date} {doc_name}{format_small}".replace('/', ' ')
        new_file_name_bytes = new_file_name.encode('utf-8')
        new_file_name = new_file_name_bytes[:250].decode('utf-8', 'ignore').rstrip(format_small).rstrip(
            format_high) + format_small
        downloads_directory = os.path.abspath(f'./Документи по ВП/{vp_num_value}')
        if os.path.isfile(os.path.join(downloads_directory, doc_type,
                                       new_file_name)):  # Документи по ВП/65892962/Інші документи/doc_name.pdf
            continue
        success_pdf_preview_switch, is_html = document_page_preview(driver, postanova, doc_name, vp_num_value)
        if success_pdf_preview_switch:
            if not os.path.exists(downloads_directory):
                os.makedirs(downloads_directory)
            delete_old_files(downloads_directory)
            if is_html:
                iframe = is_xpath_on_page(driver, '/html/body/div[1]/ui-view/section[2]/div/div[2]/div[4]/iframe')
                url = iframe.get_attribute('src')
                driver.get(url)
                with open(os.path.join(os.path.abspath(f'./Документи по ВП/{vp_num_value}'), 'document.html'),
                          "w") as f:
                    WebDriverWait(driver, 10).until(
                        ex.url_contains('blob:https')
                    )
                    f.write(driver.page_source)
                    driver.back()

            else:
                element = WebDriverWait(driver, 10).until(
                    ex.element_to_be_clickable(
                        (By.XPATH,
                         "//a[@data-ng-click='vm.events.btnViewFileDownloadFile()']"))
                )
                element.click()

            downloaded_file_name = wait_for_new_file(downloads_directory)

            rename_replace_file(downloads_directory, downloaded_file_name, new_file_name,
                                vp_num_value,
                                doc_type)
        else:
            with open(os.path.abspath('Документи по ВП/error.txt'), 'a') as f:
                f.write(f"{vp_num_value}: {doc_type} - не зберігся документ №{doc_number}\n")
            if not vp_num_in_sheet(sheet, vp_num_value):
                sheet.append((vp_num_value, secret_num))
        try:
            button_back = WebDriverWait(driver, 40).until(
                ex.element_to_be_clickable(
                    (By.XPATH, "//button[contains(text(), 'Назад')]"))
            )
            button_back.click()
        except:
            if not vp_num_in_sheet(sheet, vp_num_value):
                sheet.append((vp_num_value, secret_num))

            text = f"{vp_num_value}: {doc_type} - не збереглись документи після документу №{doc_number}\n"
            if doc_type == 'Перелік постанов':
                text = f"{vp_num_value}: {doc_type} та Інші документи - не збереглись документи після постанови №{doc_number}\n"
            with open(os.path.abspath('Документи по ВП/error.txt'), 'a') as f:
                f.write(text)
            wb.save('error_vp.xlsx')
            return
        # except Exception as err:
        #     print(err)
        #     sheet.append((vp_num_value, secret_num))
        #     with open(os.path.abspath('Документи по ВП/error.txt'), 'a') as f:
        #         f.write(f"{vp_num_value}; {doc_type} - деякі з постанов не вдалось завантажити\n")
    wb.save('error_vp.xlsx')
    return True


def create_or_take_workbook():
    exists = os.path.isfile('error_vp.xlsx')
    if not exists:
        wb_append = Workbook()
    else:
        wb_append = load_workbook('error_vp.xlsx')
    return wb_append


def vp_num_in_sheet(sheet, vp_num):
    for row in sheet.iter_rows(values_only=True):
        if vp_num in row[0]:
            return True
    return False


def other_document_if_exist(driver, text):
    try:
        element = WebDriverWait(driver, 5).until(
            ex.visibility_of_all_elements_located((By.XPATH,
                                                   f"//tr[@data-ng-repeat='item in vm.data.vpView.otherDocuments' and contains(., '{text}')]"))
        )
        return element[-1]
    except:
        return False


formats = [".pdf", ".PDF", '.zip', '.ZIP', ".html", ".HTML"]


def wait_for_new_file(downloads_directory):
    while True:
        files = os.listdir(downloads_directory)
        for file in files:
            if any(file.endswith(format_name) for format_name in formats):
                return os.path.join(downloads_directory, file)
        time.sleep(1)


def delete_old_files(directory):
    for filename in os.listdir(directory):
        if any(filename.endswith(format_name) for format_name in formats):
            filepath = os.path.join(directory, filename)
            os.remove(filepath)
            print("Видалено файл:", filepath)


def rename_replace_file(downloads_directory, downloaded_file_name, new_file_name, vp_num_value,
                        doc_type):
    downloaded_file_path = os.path.join(downloads_directory, downloaded_file_name)  # шлях скаченого файлу

    new_file_path = os.path.join(downloads_directory, new_file_name)  # шлях скаченого файлу з новим ім'ям
    os.rename(downloaded_file_path, new_file_path)
    new_directory = os.path.abspath(f'./Документи по ВП/{vp_num_value}/{doc_type}')
    if not os.path.exists(new_directory):
        os.makedirs(new_directory)
    new_file_path = os.path.join(downloads_directory, new_file_name)
    move_file(new_file_path, os.path.join(new_directory, new_file_name))


def move_file(source, destination):
    # Перевірити, чи файл існує у призначеній директорії
    if os.path.exists(destination):
        # Видалити існуючий файл
        os.remove(source)
    # Перемістити файл
    else:
        shutil.move(source, destination)


def document_page_preview(driver, document, doc_name, vp_num_value):
    # try:
    #     doc_name_to_click = WebDriverWait(document, 10).until(
    #         ex.element_to_be_clickable((By.CLASS_NAME, 'init-link'))
    #     )
    #     doc_name_to_click.click()
    #     res, is_html = wait_pdf(driver, document, doc_name, vp_num_value)
    # except Exception as err:
    #     print(err)
    #     res = None
    #     is_html = None
    # return res, is_html
    doc_name_to_click = WebDriverWait(document, 10).until(
        ex.element_to_be_clickable((By.CLASS_NAME, 'init-link'))
    )
    doc_name_to_click.click()
    res, is_html = wait_pdf(driver, document, doc_name, vp_num_value)
    return res, is_html


def search_attempts(driver, element):
    count = 0
    success_search = False
    while count != 3 and not success_search:
        try:
            success_btn = WebDriverWait(driver, 5).until(
                ex.presence_of_element_located((By.CLASS_NAME, 'btn--color-info.btn--state-successful')))
            success_search = True
        except Exception as er:
            success_search = False
            count += 1
            element.click()
    return success_search


def wait_pdf(driver: webdriver.Chrome, document, doc_name, vp_num_value):
    count = 0
    success_search = False
    is_html = False

    while count != 3:
        try:
            WebDriverWait(driver, 5).until(
                ex.visibility_of_element_located((By.XPATH, "//section[@class='print-hide' and @id='page3']")))
            div_file_types = WebDriverWait(driver, 10).until(
                ex.visibility_of_element_located((By.XPATH,
                                                  "//div[contains(@class, 'row--align-center') and contains(@data-ng-show, '!vm.ui.fileView.isLoading')]")))
            all_div_types = WebDriverWait(div_file_types, 5).until(
                ex.visibility_of_element_located((By.XPATH, "//div[@class='row row--align-center' and @ng-show]")))
            file_type = all_div_types.get_attribute("ng-show").split('==')[-1].replace("'", '').strip()

            if file_type == 'pdf':
                WebDriverWait(driver, 5).until(
                    ex.visibility_of_element_located((By.ID, "pdfPlaceholder"))
                )
                return True, False
            elif file_type == 'html':
                return True, True
            elif file_type == 'unknown':
                return True, False
        except Exception as err:
            print(err)
            count += 1
            print('pdf page loading')

    return False, False


def try_back(driver, document, doc_name, vp_num_value):
    try:
        button_back = WebDriverWait(driver, 20).until(
            ex.visibility_of_element_located(
                (By.XPATH, "//button[contains(text(), 'Назад')]"))
        )
        button_back.click()
        doc_name_to_click = WebDriverWait(document, 10).until(
            ex.element_to_be_clickable((By.CLASS_NAME, 'init-link'))
        )
        doc_name_to_click.click()
    except:
        with open(os.path.abspath('Документи по ВП/error.txt'), 'a') as f:
            f.write(f"Сторінку з {doc_name} не завантажено (vp: {vp_num_value})\n")


def is_xpath_on_page(driver, path, wait=5):
    try:
        element = WebDriverWait(driver, wait).until(
            ex.visibility_of_element_located(
                (By.XPATH, path)
            )
        )
        return element
    except:
        return False


def get_until_not_vp_num(driver: webdriver.Firefox, url):
    vp_num = None
    count = 0
    while count != 3 and not vp_num:
        count += 1
        driver.get(url)

        element = WebDriverWait(driver, 10).until(
            ex.visibility_of_element_located((By.XPATH, "//label[span='Доступ сторін до виконавчого провадження']")))
        element.click()
        print('waiting for visibility Номер ВП')
        try:
            vp_num = WebDriverWait(driver, 10).until(
                ex.visibility_of_element_located(
                    (By.CSS_SELECTOR, 'input[data-ng-model="vm.data.model.filterVPData.VpNum"]'))
            )
        except:
            print('Not vp num on page. Next try...')
            driver.refresh()
    else:
        return vp_num
