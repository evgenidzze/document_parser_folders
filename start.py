import time
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
from selenium import webdriver
from selenium_manage import perform_selenium_actions
import os


def get_vp_secret():
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    if file_path:
        os.remove('error_vp.xlsx') if os.path.isfile('error_vp.xlsx') else None
        os.remove('Документи по ВП/error.txt') if os.path.isfile('Документи по ВП/error.txt') else None

        workbook = load_workbook(file_path)
        sheet = workbook.active
        for row in sheet.iter_rows(values_only=True, min_row=2):
            start_driver(row)

        if os.path.isfile('error_vp.xlsx'):
            os.remove('Документи по ВП/error.txt') if os.path.isfile('Документи по ВП/error.txt') else None
            workbook_error = load_workbook(os.path.abspath('error_vp.xlsx'))
            sheet_error = workbook_error.active
            for row in sheet_error.iter_rows(values_only=True):
                start_driver(row)


def start_driver(row):
    vp_num_value, secret_num_value = row[0], row[1]
    options = webdriver.ChromeOptions()
    options.add_experimental_option('prefs', {
        "download.default_directory": os.path.abspath(f'./Документи по ВП/{vp_num_value}'),
    })
    driver = webdriver.Chrome(options=options)
    print('data: ', row)
    perform_selenium_actions(vp_num_value, secret_num_value, driver)
    driver.close()


root = tk.Tk()

root.title("Document Parser Folders")

file_button = tk.Button(root, text="Choose XLSX File", command=get_vp_secret)
file_button.pack(pady=20)

root.mainloop()
